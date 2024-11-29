from flask import Flask, render_template, request, redirect, url_for, flash, send_file, session, jsonify
import os
import pandas as pd
import requests
from werkzeug.utils import secure_filename
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from threading import Thread, Event
import uuid

app = Flask(__name__)
app.secret_key = 'your_secret_key'  # Replace with a secure key

# Global dictionary to store job details
jobs = {}

def get_country_name(country_code):
    """Retrieve the full country name from the country code."""
    if not country_code:
        return "Unknown Country"
    try:
        response = requests.get(f"https://restcountries.com/v3.1/alpha/{country_code}")
        if response.status_code == 200:
            country_data = response.json()
            return country_data[0]["name"]["common"]
        else:
            return "Unknown Country"
    except:
        return "Unknown Country"

def get_ip_info(api_key, ip):
    """Fetch IP reputation information from VirusTotal."""
    url = f"https://www.virustotal.com/api/v3/ip_addresses/{ip}"
    headers = {"x-apikey": api_key}
    response = requests.get(url, headers=headers)

    if response.status_code == 401:
        return None, None, None

    if response.status_code == 200:
        data = response.json()
        isp = data.get("data", {}).get("attributes", {}).get("as_owner", "Unknown ISP")
        country_code = data.get("data", {}).get("attributes", {}).get("country", "")
        country = get_country_name(country_code)
        malicious_count = data.get("data", {}).get("attributes", {}).get("last_analysis_stats", {}).get("malicious", 0)
        reputation = classify_reputation(malicious_count)
        return isp, country, reputation
    else:
        return "Error", "Error", "Error"

def classify_reputation(score):
    """Classify the IP reputation based on the malicious score."""
    if score == 0:
        return "Safe"
    elif 1 <= score <= 10:
        return "Neutral"
    elif score > 10:
        return "Poor"
    return "Unknown"

def get_hash_reputation(api_key, file_hash):
    """Fetch file hash reputation information from VirusTotal."""
    url = f"https://www.virustotal.com/api/v3/files/{file_hash}"
    headers = {"x-apikey": api_key}
    response = requests.get(url, headers=headers)

    if response.status_code == 200:
        data = response.json()
        attributes = data.get("data", {}).get("attributes", {})

        last_analysis_stats = attributes.get("last_analysis_stats", {})
        malicious_count = last_analysis_stats.get("malicious", 0)
        total_engines = sum(last_analysis_stats.values())
        reputation = classify_hash_reputation(malicious_count, total_engines)

        community_score = attributes.get("reputation", 0)

        signature_info = attributes.get("signature_info", {})
        file_signer = signature_info.get("subject", "Unknown")

        file_type = attributes.get("type_description", "Unknown")
        file_size = attributes.get("size", "Unknown")
        file_names = attributes.get("names", [])

        return reputation, malicious_count, total_engines, community_score, file_signer, file_type, file_size, file_names
    elif response.status_code == 404:
        return "Not Found", 0, 0, 0, "Unknown", "Unknown", "Unknown", []
    elif response.status_code == 401:
        return "Invalid API Key", 0, 0, 0, "Unknown", "Unknown", "Unknown", []
    else:
        return "Error", 0, 0, 0, "Unknown", "Unknown", "Unknown", []

def classify_hash_reputation(malicious_count, total_engines):
    """Classify the file hash reputation based on malicious detections."""
    if total_engines == 0:
        return "Unknown"
    
    malicious_percentage = (malicious_count / total_engines) * 100
    
    if malicious_percentage == 0:
        return "Safe"
    elif malicious_percentage <= 20:
        return "Low Risk"
    elif malicious_percentage <= 50:
        return "Moderate Risk"
    else:
        return "High Risk"

def get_comments(api_key, resource_type, resource_id, limit=5):
    """
    Fetch top community comments from VirusTotal for a given resource.

    :param api_key: Your VirusTotal API key.
    :param resource_type: Type of the resource ('ip_addresses' or 'files').
    :param resource_id: The IP address or file hash.
    :param limit: Number of comments to retrieve.
    :return: List of comments with user, date, and text.
    """
    url = f"https://www.virustotal.com/api/v3/{resource_type}/{resource_id}/comments"
    headers = {"x-apikey": api_key}
    params = {"limit": limit}
    response = requests.get(url, headers=headers, params=params)

    if response.status_code == 200:
        data = response.json()
        comments = data.get('data', [])
        top_comments = []
        for comment in comments:
            attributes = comment.get('attributes', {})
            top_comments.append({
                'user': attributes.get('user', 'Anonymous'),
                'date': attributes.get('date', ''),
                'comment': attributes.get('text', '')
            })
        return top_comments
    else:
        # You can log the error or handle it as needed
        return []

@app.route('/')
def index():
    """Home page for Single IP Lookup."""
    api_key = session.get('api_key', '')
    return render_template('index.html', api_key=api_key)

@app.route('/lookup_ip', methods=['POST'])
def lookup_ip():
    """Handle Single IP Lookup form submission."""
    ip = request.form['ip']
    api_key = request.form['api_key']

    if not api_key or not ip:
        flash('Please enter API Key and IP address.')
        return redirect(url_for('index'))

    # Save API Key in session
    session['api_key'] = api_key

    isp, country, reputation = get_ip_info(api_key, ip)
    if isp is None:
        flash('Invalid API Key.')
        return redirect(url_for('index'))

    # Fetch top 5 community comments for the IP
    comments = get_comments(api_key, 'ip_addresses', ip, limit=5)

    result = {
        'ip': ip,
        'isp': isp,
        'country': country,
        'reputation': reputation,
        'comments': comments  # Add comments to the result
    }

    return render_template('lookup_result.html', result=result)

@app.route('/bulk_upload')
def bulk_upload():
    """Page for Bulk IP Lookup file upload."""
    api_key = session.get('api_key', '')
    return render_template('bulk_upload.html', api_key=api_key)

@app.route('/process_bulk_upload', methods=['POST'])
def process_bulk_upload():
    """Handle Bulk IP Lookup file upload and start analysis."""
    file = request.files['file']
    api_key = request.form['api_key']

    if not api_key:
        flash('Please enter your VirusTotal API Key.')
        return redirect(url_for('bulk_upload'))

    # Save API Key in session
    session['api_key'] = api_key

    if file.filename == '':
        flash('No selected file.')
        return redirect(url_for('bulk_upload'))

    filename = secure_filename(file.filename)
    file_ext = os.path.splitext(filename)[1].lower()

    if file_ext not in ['.csv', '.xls', '.xlsx']:
        flash('Unsupported file format. Please upload a CSV or Excel file.')
        return redirect(url_for('bulk_upload'))

    # Save the uploaded file to disk
    upload_folder = app.config.get('UPLOAD_FOLDER', 'uploads')
    os.makedirs(upload_folder, exist_ok=True)
    file_path = os.path.join(upload_folder, filename)
    file.save(file_path)

    # Flash success message
    flash('File successfully uploaded.')

    # Generate a unique job ID
    job_id = str(uuid.uuid4())

    # Initialize job details
    jobs[job_id] = {
        'status': 'Processing',
        'progress': 0,
        'result_file': None,
        'message': 'File successfully uploaded. Starting analysis...',
        'cancel_event': Event()
    }

    # Start background thread to process the file
    thread = Thread(target=process_file_thread, args=(job_id, file_path, api_key))
    thread.start()

    # Redirect to the progress page
    return redirect(url_for('bulk_progress', job_id=job_id))

def process_file_thread(job_id, file_path, api_key):
    """Background thread function to process the uploaded file."""
    try:
        # Read the file
        file_ext = os.path.splitext(file_path)[1].lower()
        if file_ext == '.csv':
            ip_df = pd.read_csv(file_path)
        else:
            ip_df = pd.read_excel(file_path)

        # Check for acceptable column names (case-insensitive)
        acceptable_columns = ['IP', 'ip_address', 'IPAddress']
        columns_lower = [col.lower() for col in ip_df.columns]
        ip_column = None
        for col in acceptable_columns:
            if col.lower() in columns_lower:
                ip_column = ip_df.columns[columns_lower.index(col.lower())]
                break

        if not ip_column:
            jobs[job_id]['status'] = 'Error'
            jobs[job_id]['message'] = "The uploaded file must contain a column named 'IP', 'ip_address', or 'IPAddress'."
            return

        total_ips = len(ip_df)
        if total_ips == 0:
            jobs[job_id]['status'] = 'Error'
            jobs[job_id]['message'] = "The uploaded file contains no IP addresses."
            return

        output_data = []

        for idx, ip in enumerate(ip_df[ip_column]):
            # Check if cancellation has been requested
            if jobs[job_id]['cancel_event'].is_set():
                jobs[job_id]['status'] = 'Canceled'
                jobs[job_id]['message'] = 'IP analysis canceled by user.'
                return

            isp, country, reputation = get_ip_info(api_key, ip)
            if isp is None:
                jobs[job_id]['status'] = 'Error'
                jobs[job_id]['message'] = 'Invalid API Key.'
                return

            output_data.append({
                "S. No.": idx + 1,
                "IP": ip,
                "ISP": isp,
                "Country": country,
                "Reputation": reputation
            })

            # Update progress
            jobs[job_id]['progress'] = int((idx + 1) / total_ips * 100)

        output_df = pd.DataFrame(output_data)

        # Save the result to a file
        downloads_folder = app.config.get('DOWNLOAD_FOLDER', 'downloads')
        os.makedirs(downloads_folder, exist_ok=True)
        output_file = os.path.join(downloads_folder, f'{job_id}_result.xlsx')

        wb = Workbook()
        ws = wb.active
        ws.title = "Source IP Details"

        for r_idx, row in enumerate(dataframe_to_rows(output_df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                                     top=Side(style="thin"), bottom=Side(style="thin"))
                cell.border = thin_border
                if r_idx == 1:
                    cell.fill = PatternFill(start_color="16365C", end_color="16365C", fill_type="solid")
                    cell.font = Font(color="FFFFFF", bold=True)

        for column in ws.columns:
            max_length = max(len(str(cell.value)) for cell in column if cell.value)
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width

        wb.save(output_file)

        jobs[job_id]['status'] = 'Completed'
        jobs[job_id]['result_file'] = output_file
        jobs[job_id]['message'] = 'IP analysis done.'

    except Exception as e:
        jobs[job_id]['status'] = 'Error'
        jobs[job_id]['message'] = f'Error processing file: {str(e)}'

    finally:
        # Remove the uploaded file
        if os.path.exists(file_path):
            os.remove(file_path)

@app.route('/bulk_progress/<job_id>')
def bulk_progress(job_id):
    """Display the progress of the bulk IP analysis."""
    if job_id not in jobs:
        flash('Invalid job ID.')
        return redirect(url_for('bulk_upload'))

    return render_template('bulk_progress.html', job_id=job_id)

@app.route('/progress/<job_id>')
def get_progress(job_id):
    """Return the current progress and status of the job."""
    if job_id in jobs:
        return jsonify({
            'status': jobs[job_id]['status'],
            'progress': jobs[job_id]['progress'],
            'message': jobs[job_id]['message']
        })
    else:
        return jsonify({'status': 'Error', 'message': 'Invalid job ID.'})

@app.route('/download/<job_id>')
def download_result(job_id):
    """Allow the user to download the analysis result."""
    if job_id in jobs and jobs[job_id]['status'] == 'Completed':
        return send_file(jobs[job_id]['result_file'], as_attachment=True)
    else:
        flash('Result file not available.')
        return redirect(url_for('bulk_upload'))

@app.route('/cancel/<job_id>', methods=['POST'])
def cancel_job(job_id):
    """Cancel an ongoing IP analysis job."""
    if job_id in jobs and jobs[job_id]['status'] == 'Processing':
        jobs[job_id]['cancel_event'].set()
        flash('IP analysis has been canceled.')
    else:
        flash('Cannot cancel this job.')
    return redirect(url_for('bulk_progress', job_id=job_id))

@app.route('/hash_lookup')
def hash_lookup():
    """Page for Hash Reputation Check."""
    api_key = session.get('api_key', '')
    return render_template('hash_lookup.html', api_key=api_key)

@app.route('/lookup_hash', methods=['POST'])
def lookup_hash():
    """Handle Hash Reputation Check form submission."""
    file_hash = request.form['file_hash']
    api_key = request.form['api_key']

    if not api_key or not file_hash:
        flash('Please enter API Key and File Hash.')
        return redirect(url_for('hash_lookup'))

    # Save API Key in session
    session['api_key'] = api_key

    result = get_hash_reputation(api_key, file_hash)
    status = result[0]
    if status == "Unknown":
        flash('Unable to retrieve hash reputation.')
        return redirect(url_for('hash_lookup'))
    elif status == "Not Found":
        flash('File hash not found in VirusTotal database.')
        return redirect(url_for('hash_lookup'))
    elif status == "Invalid API Key":
        flash('Invalid API Key.')
        return redirect(url_for('hash_lookup'))
    elif status == "Error":
        flash('An error occurred while retrieving data.')
        return redirect(url_for('hash_lookup'))

    reputation, malicious_count, total_engines, community_score, file_signer, file_type, file_size, file_names = result

    if community_score == 0:
        community_reputation = "Safe"
    elif community_score == 1:
        community_reputation = "Suspicious"
    else:
        community_reputation = "Malicious"

    # Fetch top 5 community comments for the file hash
    comments = get_comments(api_key, 'files', file_hash, limit=5)

    result_data = {
        'file_hash': file_hash,
        'reputation': reputation,
        'community_reputation': community_reputation,
        'malicious_count': malicious_count,
        'total_engines': total_engines,
        'file_type': file_type,
        'file_size': file_size,
        'file_signer': file_signer,
        'file_names': file_names,
        'comments': comments  # Add comments to the result
    }

    return render_template('hash_result.html', result=result_data)

@app.route('/api_key')
def api_key_page():
    """Page for API Key Management."""
    api_key = session.get('api_key', '')
    return render_template('api_key.html', api_key=api_key)

@app.route('/save_api_key', methods=['POST'])
def save_api_key_route():
    """Save the API Key to the session."""
    api_key = request.form['api_key']
    if not api_key:
        flash('Please enter an API Key before saving.')
        return redirect(url_for('api_key_page'))

    # Save API Key in session
    session['api_key'] = api_key
    flash('API Key saved successfully.')
    return redirect(url_for('api_key_page'))

if __name__ == '__main__':
    app.run(debug=True)
